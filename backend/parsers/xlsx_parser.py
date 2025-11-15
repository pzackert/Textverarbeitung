"""XLSX Parser - Einfache Zellen-Extraktion"""
from pathlib import Path
from typing import Dict, Any
from openpyxl import load_workbook
from backend.parsers.base_parser import BaseParser
from backend.utils.logger import setup_logger

logger = setup_logger(__name__)


class XLSXParser(BaseParser):
    def parse(self, file_path: Path) -> Dict[str, Any]:
        try:
            metadata = self.get_metadata(file_path)
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            
            text_parts = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"=== {sheet_name} ===")
                
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        text_parts.append(" | ".join(row_values))
            
            full_text = "\n".join(text_parts)
            metadata.update({"sheet_count": len(wb.sheetnames), "sheets": wb.sheetnames})
            wb.close()
            
            logger.info(f"XLSX geparst: {file_path.name} ({len(wb.sheetnames)} Sheets)")
            return {"text": full_text, "metadata": metadata, "error": None}
            
        except Exception as e:
            logger.error(f"XLSX Parse Error: {file_path.name} - {e}")
            return {"text": "", "metadata": self.get_metadata(file_path), "error": str(e)}


if __name__ == "__main__":
    # Test
    print("XLSX Parser Test")
    
    test_dir = Path(__file__).parent.parent.parent / "data" / "input"
    test_dir.mkdir(parents=True, exist_ok=True)
    
    xlsx_files = list(test_dir.glob("*.xlsx"))
    
    if xlsx_files:
        parser = XLSXParser()
        result = parser.parse(xlsx_files[0])
        
        print(f"\n✓ Test erfolgreich!")
        print(f"  Datei: {result['metadata']['filename']}")
        print(f"  Sheets: {result['metadata'].get('sheets', 'N/A')}")
        print(f"  Zeichen: {len(result['text'])}")
        
        if result['text']:
            print(f"\n  Erste 200 Zeichen:")
            print(f"  {result['text'][:200]}...")
    else:
        print(f"  Keine XLSX-Dateien in {test_dir} gefunden")
