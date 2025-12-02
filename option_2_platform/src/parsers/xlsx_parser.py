from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
from typing import List

from .base import BaseParser
from .models import Document
from .exceptions import CorruptedFileError, EmptyDocumentError

class XlsxParser(BaseParser):
    supported_formats = ['xlsx']
    
    def parse(self, file_path: str) -> List[Document]:
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CorruptedFileError(f"Cannot open XLSX: {str(e)}")
        
        documents = []
        file_size = path.stat().st_size
        
        try:
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                
                if len(rows) < 2:
                    continue
                
                headers = [str(cell) if cell else '' for cell in rows[0]]
                
                for row_idx, row in enumerate(rows[1:], start=2):
                    if not any(cell for cell in row):
                        continue
                    
                    content_lines = []
                    for header, cell in zip(headers, row):
                        if cell is not None:
                            content_lines.append(f"{header}: {cell}")
                    
                    if content_lines:
                        content = '\n'.join(content_lines)
                        
                        metadata = {
                            "sheet_name": sheet,
                            "row_number": row_idx,
                            "column_headers": headers,
                            "file_size": file_size,
                            "modified_date": datetime.fromtimestamp(path.stat().st_mtime).isoformat()
                        }
                        
                        documents.append(Document(
                            content=content,
                            metadata=metadata,
                            source_file=str(file_path),
                            file_type="xlsx"
                        ))
        except Exception as e:
            raise CorruptedFileError(f"Error processing XLSX: {str(e)}")
        
        if not documents:
            raise EmptyDocumentError("No data rows extracted from XLSX")
        
        return documents
