import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

import fitz  # PyMuPDF
from docx import Document
from docx.enum.text import WD_COLOR_INDEX
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


class AnnotationService:
    """
    Service for creating annotated copies of documents across multiple formats.
    Supports PDF highlights, DOCX text highlighting, XLSX cell highlighting,
    and TXT sidecar metadata files when visual annotation is not possible.
    """

    COLORS = {
        "success": (0, 1, 0),
        "pass": (0, 1, 0),
        "warning": (1, 1, 0),
        "fail": (1, 0, 0),
        "default": (1, 1, 0),
    }

    def _color(self, status: str) -> tuple:
        return self.COLORS.get(status.lower(), self.COLORS["default"]) if status else self.COLORS["default"]

    def annotate_pdf(
        self,
        file_path: Path,
        text: str,
        output_path: Path,
        page: Optional[int] = None,
        status: str = "default",
    ) -> Dict[str, Any]:
        """Highlight matching text in a PDF and write a copy to output_path."""
        result = {"annotated_file": None, "highlights": []}
        if not file_path.exists():
            logger.error(f"PDF not found: {file_path}")
            return result

        doc = fitz.open(file_path)
        try:
            for idx, page_obj in enumerate(doc):
                if page is not None and idx != max(page - 1, 0):
                    continue

                instances = page_obj.search_for(text, quads=True)
                if not instances and len(text) > 80:
                    instances = page_obj.search_for(text[:80], quads=True)

                for quad in instances:
                    annot = page_obj.add_highlight_annot(quad)
                    annot.set_colors(stroke=self._color(status))
                    annot.set_opacity(0.4)
                    annot.update()
                    rect = annot.rect
                    result["highlights"].append(
                        {
                            "page": idx + 1,
                            "bbox": {
                                "x": rect.x0,
                                "y": rect.y0,
                                "width": rect.width,
                                "height": rect.height,
                            },
                        }
                    )

            output_path.parent.mkdir(parents=True, exist_ok=True)
            doc.save(output_path)
            logger.info(f"Annotated PDF saved to {output_path}")
            result["annotated_file"] = str(output_path)
            return result
        except Exception as exc:
            logger.error(f"Failed to annotate PDF {file_path}: {exc}")
            return result
        finally:
            doc.close()

    def annotate_docx(self, file_path: Path, text: str, output_path: Path) -> Dict[str, Any]:
        """Highlight matching runs in a DOCX file."""
        result = {"annotated_file": None, "highlights": 0}
        if not file_path.exists():
            logger.error(f"DOCX not found: {file_path}")
            return result

        try:
            doc = Document(file_path)
            hits = 0
            for paragraph in doc.paragraphs:
                if text in paragraph.text:
                    for run in paragraph.runs:
                        if text in run.text:
                            run.font.highlight_color = WD_COLOR_INDEX.YELLOW
                            hits += 1
            output_path.parent.mkdir(parents=True, exist_ok=True)
            doc.save(output_path)
            if hits == 0:
                logger.warning(f"No highlights created in DOCX {file_path}")
            result["annotated_file"] = str(output_path)
            result["highlights"] = hits
            return result
        except Exception as exc:
            logger.error(f"Failed to annotate DOCX {file_path}: {exc}")
            return result

    def annotate_xlsx(self, file_path: Path, cell_ref: str, output_path: Path) -> Dict[str, Any]:
        """Highlight a specific cell in an XLSX file."""
        result = {"annotated_file": None, "highlights": 0}
        if not file_path.exists():
            logger.error(f"XLSX not found: {file_path}")
            return result

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            cell = ws[cell_ref]
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            result["annotated_file"] = str(output_path)
            result["highlights"] = 1
            return result
        except Exception as exc:
            logger.error(f"Failed to annotate XLSX {file_path}: {exc}")
            return result

    def annotate_txt(
        self,
        file_path: Path,
        line_number: int,
        output_path: Path,
        text: str,
        criterion_id: str,
    ) -> Dict[str, Any]:
        """Create a sidecar metadata file for TXT references and return its path."""
        result = {"annotated_file": None, "meta_file": None, "highlights": 1}
        if not file_path.exists():
            logger.error(f"TXT not found: {file_path}")
            return result

        meta_path = Path(f"{output_path}.meta.json")
        meta = {
            "original": file_path.name,
            "references": [
                {"line": line_number, "text": text, "criterion": criterion_id}
            ],
        }
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(meta_path, "w", encoding="utf-8") as handle:
                json.dump(meta, handle, ensure_ascii=False, indent=2)
            logger.info(f"TXT meta written to {meta_path}")
            result["meta_file"] = str(meta_path)
            return result
        except Exception as exc:
            logger.error(f"Failed to write TXT meta for {file_path}: {exc}")
            return result

    def annotate_document(
        self,
        file_path: Path,
        evidence: Dict[str, Any],
        criterion_id: str,
        output_dir: Path,
        status: str = "default",
    ) -> Optional[Dict[str, Any]]:
        """Universal annotation dispatcher for PDF, DOCX, XLSX, and TXT."""
        suffix = file_path.suffix.lower()
        annotated_path = output_dir / f"{file_path.stem}_annotated{file_path.suffix}"
        reference = None
        meta_file = None
        highlights = []
        highlights_count = 0

        if suffix == ".pdf":
            pdf_result = self.annotate_pdf(
                file_path=file_path,
                text=evidence.get("text", ""),
                output_path=annotated_path,
                page=evidence.get("page"),
                status=status,
            )
            annotated_file = pdf_result.get("annotated_file")
            highlights = pdf_result.get("highlights", [])
            highlights_count = len(highlights)
            reference = f"Seite {evidence.get('page', '?')}" if evidence.get("page") else "Seite ?"
        elif suffix == ".docx":
            docx_result = self.annotate_docx(file_path, evidence.get("text", ""), annotated_path)
            annotated_file = docx_result.get("annotated_file")
            highlights_count = docx_result.get("highlights", 0)
            reference = evidence.get("reference") or "Absatz"
        elif suffix == ".xlsx":
            cell_ref = evidence.get("cell") or evidence.get("cell_ref") or evidence.get("reference") or "A1"
            xlsx_result = self.annotate_xlsx(file_path, cell_ref, annotated_path)
            annotated_file = xlsx_result.get("annotated_file")
            highlights_count = xlsx_result.get("highlights", 0)
            reference = f"Zelle {cell_ref}"
        elif suffix == ".txt":
            txt_result = self.annotate_txt(
                file_path=file_path,
                line_number=evidence.get("line", 1),
                output_path=annotated_path,
                text=evidence.get("text", ""),
                criterion_id=criterion_id,
            )
            meta_file = txt_result.get("meta_file")
            annotated_file = txt_result.get("annotated_file")
            highlights_count = txt_result.get("highlights", 1)
            reference = f"Zeile {evidence.get('line', '?')}"
        else:
            logger.warning(f"Unsupported format for annotation: {suffix}")
            return None

        return {
            "annotated_file": annotated_file,
            "reference": reference,
            "format": suffix.lstrip("."),
            "meta_file": meta_file,
            "original": str(file_path),
            "criterion_id": criterion_id,
            "highlights": highlights,
            "highlights_count": highlights_count,
            "text": evidence.get("text"),
        }

    def create_annotated_pdf(
        self,
        input_path: Path,
        output_path: Path,
        citations: List[Dict[str, Any]],
        status: str = "default",
    ) -> bool:
        """Backward-compatible PDF citation highlighter."""
        if not input_path.exists():
            logger.error(f"Input PDF not found: {input_path}")
            return False

        try:
            doc = fitz.open(input_path)
            color = self._color(status)
            for citation in citations:
                self._highlight_citation(doc, citation, color)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            doc.save(output_path)
            doc.close()
            logger.info(f"Successfully created annotated PDF: {output_path}")
            return True
        except Exception as exc:
            logger.error(f"Failed to annotate PDF {input_path}: {exc}")
            return False

    def _highlight_citation(self, doc, citation: Dict[str, Any], color: tuple):
        try:
            page_num = citation.get("page", 1) - 1
            if page_num < 0 or page_num >= len(doc):
                logger.warning(f"Page {page_num + 1} out of range for document")
                return

            page = doc[page_num]
            text_to_find = (
                citation.get("quote")
                or citation.get("text_segment")
                or citation.get("text")
                or citation.get("content")
            )

            if not text_to_find:
                logger.warning("No text found in citation to highlight")
                return

            instances = page.search_for(text_to_find, quads=True)
            if not instances and len(text_to_find) > 100:
                instances = page.search_for(text_to_find[:100], quads=True)

            for quad in instances:
                annot = page.add_highlight_annot(quad)
                annot.set_colors(stroke=color)
                annot.set_opacity(0.4)
                annot.update()

            if instances:
                logger.debug(f"Highlighted {len(instances)} instance(s) on page {page_num + 1}")
        except Exception as exc:
            logger.warning(f"Could not highlight citation on page {citation.get('page')}: {exc}")

    def annotate_from_rag_results(
        self,
        project_id: str,
        original_filename: str,
        rag_sources: List[Dict[str, Any]],
        status: str = "default",
    ) -> Optional[str]:
        """Compatibility helper: annotate PDF from RAG sources."""
        try:
            base_input = Path("data/input") / project_id / "uploads"
            input_path = base_input / original_filename
            if not input_path.exists():
                logger.error(f"Original file not found: {original_filename}")
                return None

            name_parts = original_filename.rsplit(".", 1)
            annotated_filename = f"{name_parts[0]}_annotated.{name_parts[1]}" if len(name_parts) == 2 else f"{original_filename}_annotated"
            output_dir = Path("data/input") / project_id / "annotated"
            output_path = output_dir / annotated_filename

            citations = []
            for source in rag_sources:
                citations.append(
                    {
                        "page": source.get("page", 1),
                        "text": source.get("text")
                        or source.get("content")
                        or source.get("page_content", ""),
                    }
                )

            success = self.create_annotated_pdf(input_path, output_path, citations, status)
            return annotated_filename if success else None
        except Exception as exc:
            logger.error(f"Failed to create annotated PDF from RAG results: {exc}")
            return None


annotation_service = AnnotationService()
