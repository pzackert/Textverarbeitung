import logging
import re
import json
from datetime import datetime, timezone
from time import perf_counter
from pathlib import Path
from typing import Any, Dict, List, Optional

import fitz
from docx import Document
from openpyxl import load_workbook

from src.core.models import Project
from src.services.annotation_service import annotation_service
from src.services.criteria_service import criteria_service
from src.services.project_service import project_service
from src.services.criteria_results_store import save_criterion_result
from src.api.dependencies import get_llm_chain, get_config

logger = logging.getLogger(__name__)


class ValidationService:
    def __init__(self):
        self.annotation_service = annotation_service

    def evaluate_criterion(self, project_id: str, criterion_id: str, llm_chain=None) -> Dict[str, Any]:
        project = project_service.get_project(project_id)
        if not project:
            raise ValueError(f"Project {project_id} not found")

        criterion = criteria_service.get_by_id(criterion_id)
        if not criterion:
            raise ValueError(f"Criterion {criterion_id} not found")

        llm_chain = llm_chain or get_llm_chain()
        start = perf_counter()
        rag_evidence = self._llm_eval(project_id, criterion, llm_chain)
        status = rag_evidence["status"]
        reason = rag_evidence["reason"]
        evidence_list = rag_evidence.get("evidence_raw", [])
        parsed_payload = rag_evidence.get("parsed", {})

        output_dir = Path("data/input") / project_id / "annotated"
        annotations: List[Dict[str, Any]] = []
        evidence_records: List[Dict[str, Any]] = []
        for evidence in evidence_list:
            file_path = Path("data/input") / project_id / "uploads" / evidence.get("dokument")
            if not file_path.exists():
                # still record evidence without annotation
                evidence_records.append(self._build_evidence_entry({
                    "filename": evidence.get("dokument"),
                    "reference": evidence.get("referenz"),
                    "path": str(file_path),
                    "text": evidence.get("text_snippet"),
                    "page": evidence.get("page"),
                    "cell": evidence.get("cell"),
                }, {}))
                continue
            annotated = self.annotation_service.annotate_document(
                file_path=file_path,
                evidence={
                    "text": evidence.get("text_snippet") or evidence.get("referenz"),
                    "page": evidence.get("page"),
                    "cell": evidence.get("cell"),
                    "reference": evidence.get("referenz"),
                },
                criterion_id=criterion.id,
                output_dir=output_dir,
                status=status,
            )
            if annotated:
                annotated["document"] = evidence.get("dokument")
                annotated["criterion_id"] = criterion.id
                annotated["status"] = status
                annotations.append(annotated)
                evidence_records.append(self._build_evidence_entry({
                    "filename": evidence.get("dokument"),
                    "reference": evidence.get("referenz"),
                    "path": str(file_path),
                    "text": evidence.get("text_snippet"),
                    "page": evidence.get("page"),
                    "cell": evidence.get("cell"),
                }, annotated))
            else:
                evidence_records.append(self._build_evidence_entry({
                    "filename": evidence.get("dokument"),
                    "reference": evidence.get("referenz"),
                    "path": str(file_path),
                    "text": evidence.get("text_snippet"),
                    "page": evidence.get("page"),
                    "cell": evidence.get("cell"),
                }, {}))

        annotated_file = annotations[0].get("annotated_file") if annotations else None
        evaluated_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        duration_sec = round(perf_counter() - start, 3)

        result = {
            "criterion_id": criterion.id,
            "criterion_name": criterion.name,
            "status": status,
            "score": rag_evidence.get("score", 0.0),
            "reason": reason,
            "begruendung": reason,
            "dokument": parsed_payload.get("dokument"),
            "referenz": parsed_payload.get("referenz"),
            "annotations": annotations,
            "annotated_file": annotated_file,
            "evaluated_at": evaluated_at,
            "evaluated_by": "system",
            "evaluation_duration_sec": duration_sec,
            "evidence": evidence_records,
        }

        project.validation_results = project.validation_results or {}
        project.validation_results[criterion.id] = result
        project.annotated_documents = project.annotated_documents or {}
        for ann in annotations:
            if ann.get("annotated_file"):
                project.annotated_documents[ann["document"]] = ann["annotated_file"]
        project_service.update_project(project)

        save_criterion_result(project_id, self._criterion_result_for_store(result))

        return result

    def evaluate_all(self, project_id: str) -> Dict[str, Any]:
        project = project_service.get_project(project_id)
        if not project:
            raise ValueError(f"Project {project_id} not found")

        criteria = criteria_service.get_all()
        results = [self.evaluate_criterion(project_id, c.id) for c in criteria]
        return {
            "project_id": project_id,
            "status": "completed",
            "criteria": results,
        }

    def _collect_evidence(self, project: Project, criterion) -> List[Dict[str, Any]]:
        matches: List[Dict[str, Any]] = []
        for doc in project.documents:
            path = Path(doc.path)
            if not path.exists():
                continue

            suffix = path.suffix.lower()
            evidence: Optional[Dict[str, Any]] = None
            if suffix == ".pdf":
                evidence = self._scan_pdf(path, criterion)
            elif suffix == ".docx":
                evidence = self._scan_docx(path, criterion)
            elif suffix == ".xlsx":
                evidence = self._scan_xlsx(path, criterion)
            elif suffix == ".txt":
                evidence = self._scan_txt(path, criterion)

            if evidence:
                evidence.update({
                    "path": str(path),
                    "document_id": doc.id,
                    "filename": doc.filename,
                })
                matches.append(evidence)

        return matches

    def _normalize_status(self, status: str) -> str:
        value = (status or "").lower()
        if value in {"green", "gruen", "grün", "ok", "success"}:
            return "grün"
        if value in {"yellow", "warn", "warning", "amber"}:
            return "gelb"
        return "rot"

    def _llm_eval(self, project_id: str, criterion, llm_chain) -> Dict[str, Any]:
        cfg = get_config()
        system_prompt = getattr(cfg.prompts, "kriterien_pruefung", None)
        answer_guideline = getattr(cfg.prompts, "antwort_richtlinie", None)
        user_prompt = criterion.prompt or criterion.lang or criterion.kurz or criterion.name

        def _strip_json_artifacts(txt: str) -> str:
            cleaned = (txt or "").strip()
            if cleaned.startswith("```"):
                cleaned = cleaned[3:]
                if cleaned.lstrip().lower().startswith("json"):
                    cleaned = cleaned.lstrip()[4:].lstrip()
                cleaned = cleaned.split("```", 1)[0]

            if "{" in cleaned and "}" in cleaned:
                cleaned = cleaned[cleaned.find("{") : cleaned.rfind("}") + 1]
            return cleaned.strip()

        def _parse_json(txt: str) -> Optional[Dict[str, Any]]:
            cleaned = _strip_json_artifacts(txt)
            if not cleaned:
                return None
            try:
                return json.loads(cleaned)
            except Exception:
                return None

        # Build strict instruction for JSON output
        json_schema_hint = (
            "Antworte NUR mit einem JSON-Objekt im Format: "
            '{"status": "rot|gelb|grün", "begruendung": "max 160 Zeichen", '
            '"dokument": "Dateiname", "referenz": "Seite X, Absatz Y"}'
        )
        combined_prompt = (
            f"{user_prompt}\n\n"
            "Erinnere dich: Antworte ausschließlich im JSON-Format ohne Markdown oder Zusatztexte. "
            f"{json_schema_hint}"
        )

        logger.info("LLM prompt for %s: %s", criterion.id, combined_prompt)

        try:
            res = llm_chain.query(
                question=combined_prompt,
                metadata_filter={"project_id": project_id, "include_global": True},
                system_prompt=system_prompt,
                answer_guideline=answer_guideline,
            )
        except TypeError:
            # Fallback for DummyLLM in tests without answer_guideline signature
            res = llm_chain.query(
                question=combined_prompt,
                metadata_filter={"project_id": project_id, "include_global": True},
                system_prompt=system_prompt,
            )
        answer = res.get("answer", "")
        logger.info("LLM raw answer try1 for %s: %s", criterion.id, answer)
        parsed = _parse_json(answer)
        retry = False

        if not parsed:
            retry = True
            retry_prompt = (
                f"{user_prompt}\n\n"
                "ACHTUNG: Vorherige Antwort war nicht valides JSON."
                " Antworte ausschließlich mit JSON ohne Markdown oder Erklärungen.\n"
                '{"status": "rot|gelb|grün", "begruendung": "max 160 Zeichen", '
                '"dokument": "Dateiname", "referenz": "Seite X, Absatz Y"}'
            )
            try:
                res = llm_chain.query(
                    question=retry_prompt,
                    metadata_filter={"project_id": project_id, "include_global": True},
                    system_prompt=system_prompt,
                    answer_guideline=answer_guideline,
                )
            except TypeError:
                res = llm_chain.query(
                    question=retry_prompt,
                    metadata_filter={"project_id": project_id, "include_global": True},
                    system_prompt=system_prompt,
                )
            answer = res.get("answer", "")
            logger.info("LLM raw answer try2 for %s: %s", criterion.id, answer)
            parsed = _parse_json(answer)

        status_raw = None
        reason = None
        dokument = None
        referenz = None
        if parsed:
            status_raw = parsed.get("status")
            reason = parsed.get("begruendung") or parsed.get("begründung") or parsed.get("begruen dung")
            dokument = parsed.get("dokument")
            referenz = parsed.get("referenz")

        status = self._normalize_status(status_raw or "gelb")

        errors: List[str] = []
        if not parsed:
            errors.append("Antwort nicht parsebar")
        if status not in {"rot", "gelb", "grün"}:
            errors.append("Status ungültig")
            status = "gelb"
        if reason:
            reason = str(reason)[:160]
        if not reason:
            reason = "Parsing Error" if parsed is None else "Keine Begründung geliefert"

        if errors and retry and not parsed:
            status = "gelb"
            reason = "Parsing Error"

        evidence_raw: List[Dict[str, Any]] = []
        for cit in res.get("citations", []):
            if hasattr(cit, "doc_name"):
                evidence_raw.append(
                    {
                        "dokument": cit.doc_name,
                        "referenz": f"Seite {cit.page}" if getattr(cit, "page", None) else None,
                        "text_snippet": cit.text_snippet,
                        "page": getattr(cit, "page", None),
                        "cell": getattr(cit, "cell", None),
                        "chunk_id": getattr(cit, "chunk_id", None),
                        "score": getattr(cit, "score", None),
                    }
                )
            elif isinstance(cit, dict):
                meta = cit.get("metadata", {}) if isinstance(cit.get("metadata", {}), dict) else {}
                doc_name = (
                    cit.get("doc_name")
                    or cit.get("document")
                    or cit.get("source")
                    or meta.get("doc_name")
                    or meta.get("source")
                )
                page_val = cit.get("page") or meta.get("page") or meta.get("page_number")
                cell_val = cit.get("cell") or meta.get("cell")
                reference = cit.get("referenz") or meta.get("referenz")
                if not reference:
                    if page_val:
                        reference = f"Seite {page_val}"
                    elif cell_val:
                        reference = f"Zelle {cell_val}"
                evidence_raw.append(
                    {
                        "dokument": doc_name,
                        "referenz": reference,
                        "text_snippet": cit.get("text_snippet") or cit.get("content") or meta.get("text"),
                        "page": page_val,
                        "cell": cell_val,
                        "chunk_id": cit.get("chunk_id") or meta.get("chunk_id"),
                        "score": cit.get("score") or meta.get("score"),
                    }
                )

        if not evidence_raw and dokument:
            evidence_raw.append(
                {
                    "dokument": dokument,
                    "referenz": referenz,
                    "text_snippet": reason or "",
                }
            )

        score = 1.0 if status == "grün" else 0.5 if status == "gelb" else 0.0

        return {
            "status": status,
            "reason": reason or "",
            "evidence_raw": evidence_raw,
            "score": score,
            "parsed": parsed or {},
        }

    def _build_evidence_entry(self, evidence: Dict[str, Any], annotated: Dict[str, Any]) -> Dict[str, Any]:
        annotated_file = annotated.get("annotated_file") or annotated.get("meta_file") if annotated else None
        annotated_name = Path(annotated_file).name if annotated_file else None
        text_snippet = (evidence.get("text") or evidence.get("reference") or "")[:200]
        reference = None
        if annotated:
            reference = annotated.get("reference")
        if not reference:
            reference = evidence.get("reference")
        if not reference:
            if evidence.get("page"):
                reference = f"Seite {evidence['page']}"
            elif evidence.get("cell"):
                reference = f"Zelle {evidence['cell']}"
            elif evidence.get("line"):
                reference = f"Zeile {evidence['line']}"

        return {
            "dokument": evidence.get("filename"),
            "dokument_original_path": f"/uploads/{evidence.get('filename')}",
            "referenz": reference,
            "page": evidence.get("page"), # Fix: Send raw page int
            "text_snippet": text_snippet,
            "annotated_file": annotated_name,
            "annotated_file_path": f"/annotated/{annotated_name}" if annotated_name else None,
        }

    def _criterion_result_for_store(self, result: Dict[str, Any]) -> Dict[str, Any]:
        allowed_keys = {
            "criterion_id",
            "criterion_name",
            "status",
            "score",
            "reason",
            "begruendung",
            "dokument",
            "referenz",
            "annotated_file",
            "evaluated_at",
            "evaluated_by",
            "evaluation_duration_sec",
            "evidence",
        }
        return {k: v for k, v in result.items() if k in allowed_keys}

    def _criterion_mentions_hamburg(self, criterion) -> bool:
        text = " ".join(
            [
                criterion.id,
                getattr(criterion, "name", ""),
                getattr(criterion, "kurz", ""),
                getattr(criterion, "lang", ""),
                getattr(criterion, "prompt", "") or "",
            ]
        ).lower()
        return "hamburg" in text or "plz" in text or criterion.id.lower() in {"k001", "k_test_auto"}

    def _extract_hamburg_match(self, text: str) -> Optional[re.Match]:
        return re.search(r"\b2[01]\d{3}\b", text)

    def _scan_pdf(self, path: Path, criterion) -> Optional[Dict[str, Any]]:
        if not self._criterion_mentions_hamburg(criterion):
            return None
        try:
            doc = fitz.open(path)
            for idx, page in enumerate(doc):
                content = page.get_text("text")
                match = self._extract_hamburg_match(content)
                if match:
                    doc.close()
                    return {"text": match.group(0), "page": idx + 1}
            doc.close()
        except Exception as exc:
            logger.warning(f"PDF scan failed for {path}: {exc}")
        return None

    def _scan_docx(self, path: Path, criterion) -> Optional[Dict[str, Any]]:
        if not self._criterion_mentions_hamburg(criterion):
            return None
        try:
            doc = Document(path)
            for paragraph in doc.paragraphs:
                match = self._extract_hamburg_match(paragraph.text)
                if match:
                    return {"text": match.group(0), "reference": paragraph.text}
        except Exception as exc:
            logger.warning(f"DOCX scan failed for {path}: {exc}")
        return None

    def _scan_xlsx(self, path: Path, criterion) -> Optional[Dict[str, Any]]:
        if not self._criterion_mentions_hamburg(criterion):
            return None
        try:
            wb = load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value and isinstance(value, str):
                        match = self._extract_hamburg_match(value)
                        if match:
                            return {
                                "text": match.group(0),
                                "cell": cell.coordinate,
                                "reference": value,
                            }
            wb.close()
        except Exception as exc:
            logger.warning(f"XLSX scan failed for {path}: {exc}")
        return None

    def _scan_txt(self, path: Path, criterion) -> Optional[Dict[str, Any]]:
        if not self._criterion_mentions_hamburg(criterion):
            return None
        try:
            with open(path, "r", encoding="utf-8") as handle:
                for idx, line in enumerate(handle, start=1):
                    match = self._extract_hamburg_match(line)
                    if match:
                        return {"text": match.group(0), "line": idx}
        except Exception as exc:
            logger.warning(f"TXT scan failed for {path}: {exc}")
        return None


validation_service = ValidationService()
