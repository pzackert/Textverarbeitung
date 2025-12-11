import os
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors

# Configuration
PROJECT_ID = "8209d44a-bfd9-42a0-a48a-a90038db444c" # Target Project ID
BASE_DIR = Path("/Users/patrick.zackert/projects/masterprojekt/option_2_platform")
DATA_DIR = BASE_DIR / "data/input" / PROJECT_ID

def ensure_directory():
    if not DATA_DIR.exists():
        print(f"Creating directory: {DATA_DIR}")
        DATA_DIR.mkdir(parents=True, exist_ok=True)
    else:
        print(f"Directory exists: {DATA_DIR}")

def create_annotated_excel():
    filename = "Businessplan_annotiert.xlsx"
    filepath = DATA_DIR / filename
    
    # Create simple excel if not exists, or just exact one
    wb = Workbook()
    ws = wb.active
    ws.title = "Businessplan"
    
    # Fill some dummy data
    ws['A1'] = "Businessplan - Annotiert"
    ws['B8'] = "Wichtiger Umsatz"
    ws['C23'] = "Kostenstelle X"
    
    # Apply Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # B8
    ws['B8'].fill = yellow_fill
    print(f"Highlighted B8 in {filename}")
    
    # C23
    ws['C23'].fill = yellow_fill
    print(f"Highlighted C23 in {filename}")
    
    wb.save(filepath)
    print(f"Saved {filepath}")

def create_annotated_pdf():
    filename = "Projektskizze_annotiert.pdf"
    filepath = DATA_DIR / filename
    
    c = canvas.Canvas(str(filepath), pagesize=A4)
    width, height = A4
    
    c.drawString(100, height - 100, "Projektskizze - Annotierte Version")
    c.drawString(100, height - 130, "Abschnitt 1: Einleitung")
    c.drawString(100, height - 150, "Dies ist der normale Text.")
    
    # Highlight Section 3
    # Draw a yellow rectangle underneath the text
    c.setFillColor(colors.yellow)
    c.rect(90, height - 300, 400, 50, fill=True, stroke=False)
    
    c.setFillColor(colors.black)
    c.drawString(100, height - 260, "Abschnitt 3: Finanzierung (Markiert)")
    c.drawString(100, height - 280, "Dieser Bereich ist kritisch und wurde markiert.")
    
    c.save()
    print(f"Saved {filepath}")

def main():
    ensure_directory()
    create_annotated_excel()
    create_annotated_pdf()
    print("Done generation.")

if __name__ == "__main__":
    main()
